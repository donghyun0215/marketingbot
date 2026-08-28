"""
코드프레소 성과 리포트(xlsx) → Supabase 임포트.
- 원본 xlsx는 대외비. data/ 에 두고 git에는 절대 올리지 않는다.
- 인바운드 문의의 실제 고객사명은 익명화하여 저장한다 (원본 파일 지시사항).
사용: SUPABASE_URL=... SUPABASE_SERVICE_ROLE_KEY=... python3 scripts/import_performance.py data/performance.xlsx
"""
import os, sys, json, re, urllib.request
import openpyxl

URL = os.environ["SUPABASE_URL"].rstrip("/")
KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
PATH = sys.argv[1] if len(sys.argv) > 1 else "data/performance.xlsx"
MEASURED = ("2026-02-27", "2026-08-25")

# 실명 → 익명 별칭. 실제 고객사명은 DB·레포 어디에도 남기지 않는다.
ANON = {}
def alias(name: str, industry_guess: str) -> str:
    if name not in ANON:
        ANON[name] = f"{industry_guess} {chr(ord('A') + len(ANON))}"
    return ANON[name]

def post(table, rows):
    if not rows:
        return
    # PostgREST는 한 배치 안의 모든 객체가 동일한 키 집합을 갖기를 요구한다.
    # 블로그 행과 링크드인 행은 컬럼이 다르므로 합집합으로 정규화한다.
    keys = sorted({k for r in rows for k in r})
    rows = [{k: r.get(k) for k in keys} for r in rows]
    req = urllib.request.Request(
        f"{URL}/rest/v1/{table}",
        data=json.dumps(rows).encode(),
        headers={"apikey": KEY, "Authorization": f"Bearer {KEY}",
                 "Content-Type": "application/json", "Prefer": "return=minimal"},
        method="POST")
    with urllib.request.urlopen(req) as r:
        print(f"  {table}: {len(rows)} rows -> {r.status}")

def num(v):
    if v in (None, "", "-"):
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None

def i(v):
    n = num(v)
    return int(n) if n is not None else None

def d(v):
    if v is None:
        return None
    s = str(v)[:10]
    return s if re.match(r"^\d{4}-\d{2}-\d{2}$", s) else None

def header_row(ws, first_col_label):
    """요약 블록을 건너뛰고 실제 표 헤더 행 번호를 찾는다."""
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(c == first_col_label for c in row if c):
            return idx
    raise RuntimeError(f"header '{first_col_label}' not found")

def table_rows(ws, first_col_label):
    hdr = header_row(ws, first_col_label)
    rows = list(ws.iter_rows(values_only=True))
    head = [c for c in rows[hdr - 1]]
    start = [n for n, c in enumerate(head) if c == first_col_label][0]
    out = []
    for r in rows[hdr:]:
        cells = list(r)[start:]
        if not cells or cells[0] in (None, ""):
            continue
        if str(cells[0]).startswith("※"):
            continue
        out.append(cells)
    return out

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)

# --- 블로그 콘텐츠 성과 (게시일, 제목, 링크, 클릭, 노출, CTR, 평균순위)
blog = []
for c in table_rows(wb["블로그_콘텐츠성과"], "게시일"):
    blog.append({"channel": "blog", "published_at": d(c[0]), "external_title": c[1],
                 "clicks": i(c[3]), "impressions": i(c[4]), "ctr": num(c[5]),
                 "avg_position": num(c[6]),
                 "measured_from": MEASURED[0], "measured_to": MEASURED[1]})

# --- 링크드인 (게시일, 미리보기, 링크, 노출, 고유노출, 클릭, CTR, 좋아요, 댓글, 공유, 참여율)
li = []
for c in table_rows(wb["링크드인_포스트성과"], "게시일"):
    li.append({"channel": "linkedin", "published_at": d(c[0]), "external_title": c[1],
               "impressions": i(c[3]), "unique_impressions": i(c[4]), "clicks": i(c[5]),
               "ctr": num(c[6]), "likes": i(c[7]), "comments": i(c[8]), "shares": i(c[9]),
               "engagement_rate": num(c[10]),
               "measured_from": MEASURED[0], "measured_to": MEASURED[1]})

# --- 유입 검색어 (게시일, 제목, 링크, 검색어, 클릭, 노출, CTR, 순위)
queries = []
for c in table_rows(wb["블로그_유입검색어"], "게시일"):
    queries.append({"external_title": c[1], "query": c[3], "clicks": i(c[4]),
                    "impressions": i(c[5]), "ctr": num(c[6]), "position": num(c[7])})

# --- 인바운드 문의 (문의일, 회사명, 직급, 문의유형, 관심서비스, 유입경로, 콘텐츠연결, 추정콘텐츠, 근거, 설명)
ATTR = {"확정": "confirmed", "추정(근거 있음)": "inferred", "특정 불가": "unknown"}
inq = []
for c in table_rows(wb["인바운드_문의"], "문의일"):
    company = str(c[1] or "")
    interest = str(c[4] or "")
    hint = "공공·연구" if any(k in company for k in ("연구원", "진흥", "공사")) else "기업"
    inq.append({"inquired_on": d(c[0]), "company_alias": alias(company, hint),
                "industry_hint": hint, "inquiry_type": c[3], "interest": interest or None,
                "source_channel": c[5], "attribution": ATTR.get(str(c[6]).strip(), "unknown"),
                "legacy_content_title": c[7] if c[7] not in (None, "-") else None,
                "note": c[9]})

print(f"parsed: blog={len(blog)} linkedin={len(li)} queries={len(queries)} inquiries={len(inq)}")
post("performance_metrics", blog + li)
post("search_queries", queries)
post("inquiries", inq)
print("anonymized companies:", len(ANON), "(실명은 저장하지 않음)")

# NOTE (재실행 시): 이 스크립트는 append-only라 그대로 다시 돌리면 행이 중복된다.
# 재임포트 전 반드시 대상 테이블을 비울 것.
#   DELETE /rest/v1/performance_metrics?id=gt.0
#   DELETE /rest/v1/search_queries?id=gt.0
#   DELETE /rest/v1/inquiries?id=gt.0
